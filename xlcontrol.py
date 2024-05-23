# OpenPyXl
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from collections import Counter

token = "xlqt" # token after every sheets name to avoid writting over anything else
wb = None
ws = None
rs = None
ms = None
excel_file_name = None #"Retting.xlsx"

def load(excel_file):
    global wb, ws, rs, ms, excel_file_name
    excel_file_name = excel_file

    # ws : active sheet, rs : result sheet, ms : mistakes sheet
    wb = load_workbook(excel_file_name)
    ws = wb[wb.sheetnames[0]]

    if "student_mistakes_"+token not in wb.sheetnames:
        rs = wb.create_sheet("student_mistakes_"+token)
        rs["A1"]="Student ID"
        rs["B1"]="Mistake(s)"
        wb.save(excel_file_name)
    else:
        rs = wb["student_mistakes_"+token]

    if "mistakes_"+token not in wb.sheetnames:
        ms = wb.create_sheet("mistakes_"+token)
        ms["A1"]="Mistake ID"
        ms["B1"]="Task ID"
        ms["C1"]="Subtask ID"
        ms["D1"]="Malus"
        ms["E1"]="Description"
        wb.save(excel_file_name)
    else:
        ms = wb["mistakes_"+token]

def extract_max_points(task):
    row1 = [cell[0] for cell in ws.iter_cols(min_row=1, max_row=1, min_col=2, values_only=True)]
    row3 = [cell[0] for cell in ws.iter_cols(min_row=3, max_row=3, min_col=2, values_only=True)]

    exercises = []

    exercise = []

    for i, value in enumerate(row3):
        if value == "a":
            if exercise:
                exercises.append(exercise)
                exercise = []
        if value is not None:
            exercise.append(row1[i])

    # Ajouter le dernier exercice à la liste des exercices
    if exercise:
        exercises.append(exercise)

    return exercises[task]


def student_get_mistakes(student_id):
    mistakes=[]
    for row in rs.iter_rows(min_row=2, max_row=rs.max_row, min_col=1, max_col=rs.max_column):
        if row[0].value == student_id:
            mistakes.append(row[1].value)
    return mistakes

def student_add_mistakes(student_id, mistake_id):
    rs.append([student_id, mistake_id])
    wb.save(excel_file_name)
   
def student_rem_mistakes(student_id, mistake_id):
    rows_to_delete = []
    for row in rs.iter_rows(min_row=1, max_row=rs.max_row):
        if row[0].value == student_id and row[1].value == mistake_id:
            rows_to_delete.append(row[0].row)
    for row_index in sorted(rows_to_delete, reverse=True):
        rs.delete_rows(row_index)
    wb.save(excel_file_name)

def get_mistakes_list():
    rng = ms.iter_rows(min_row=2, max_row=ms.max_row, min_col=1, max_col=ms.max_column)
    mistakes = []
    for m in rng:
        if any(m):
            mistakes.append({ "mistakeID" :m[0].value, "task" : m[1].value, "subtask" : m[2].value, "index" : 0, "malus" : m[3].value, "description" : m[4].value })
    return mistakes


def add_mistakes(task, subtask):
    max_value_col_A = -1

    for cell in ms['A'][1:]:
        if cell.value is not None:
            if max_value_col_A is None or int(cell.value) > max_value_col_A:
                max_value_col_A = cell.value
    ms.append([max_value_col_A+1, task, subtask])
    wb.save(excel_file_name)
    return max_value_col_A+1 # return the ID of the new mistakes


#to update mistakes
def up_mistakes(mist_id, task, subtask, malus, description):
    rng = ms.iter_rows(min_row=2, max_row=ms.max_row, min_col=1, max_col=ws.max_column)

    for row in rng:
        if row[0].value == mist_id:
            if task > 0 :
                row[1].value = task
            if subtask != "" :
                row[2].value = subtask
            if malus != 0:
                row[3].value = malus
            if description != "":
                row[4].value = description
    wb.save(excel_file_name)

def del_mistakes(mist_id):
    rng = ms.iter_rows(min_row=2, max_row=ms.max_row, min_col=1, max_col=ws.max_column)

    for row in rng:
        if row[0].value == mist_id:
            row[1].value = ""
            row[2].value = ""
            row[3].value = ""
            row[4].value = ""
    wb.save(excel_file_name)

def mistakesnumber():
    return ms.max_row-1

# Returns tasks numbers in a list
def find_task_numbers():
    rng = ws.iter_cols(min_row=2, max_row=2, min_col=2, max_col=ws.max_column)
    tasks = []
    for i in rng:
        for j in i:
            selected_cell = j
            if j.value is not None:
                tasks.append(j.value)

    return tasks

def total_student_number():
    rng = ws.iter_cols(min_row=4, max_row=ws.max_row, min_col=1, max_col=1)
    student_nbr = 0
    for i in rng:
        for j in i:
            selected_cell = j
            if j.value is not None:
                student_nbr+=1
    return student_nbr

def candidate_nbr(student_index):
    row = student_index+4
    return ws["A"+str(row)].value


def list_subtasks():
    rng = ws.iter_cols(min_row=3, max_row=3, min_col=2, max_col=ws.max_column)
    tasks = []
    for i in rng:
        for j in i:
            tasks.append(j.value)

    tasks = list(filter(lambda x: x is not None, tasks))

    return tasks


def organize_subtasks(subtasks):
    result = []
    sublist = []
    for element in subtasks:
        if element == 'a':
            if sublist:
                result.append(sublist)
            sublist = []
        sublist.append(element)
    result.append(sublist)
    return result


# Return nearby cell (left, right, up, down)
def nearby_cell(selected_cell, direction):
    if direction == 'left':
        next_cell = ws.cell(row=selected_cell.row, column=selected_cell.column - 1)
        return next_cell
    if direction == 'right':
        next_cell = ws.cell(row=selected_cell.row, column=selected_cell.column + 1)
        return next_cell

    if direction == 'up':
        next_cell = ws.cell(row=selected_cell.row - 1, column=selected_cell.column)
        return next_cell

    if direction == 'down':
        next_cell = ws.cell(row=selected_cell.row + 1, column=selected_cell.column)
        return next_cell

def get_column_b_occurrences(mistakes):
    column_b_values = [cell.value for row in rs.iter_rows(min_row=2, min_col=2, max_col=2) for cell in row if cell.value in mistakes]

    return Counter(column_b_values)

# Check
#print(organize_subtasks(list_subtasks())[0])
